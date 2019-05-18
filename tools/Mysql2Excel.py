#!/usr/bin/env python
# @Time    : 2019/5/18 20:34
# @Author  : LiuWei
# @Site    : 
# @File    : Mysql2Excel.py
# @Software: PyCharm
# -*- coding: utf-8 -*-
"""
功能简述：读取MySQL数据，转到Excel文件
需要的安装：
pip3 install openpyxl
pip3 install PyMySQL

参考链接：
Python3 MySQL 数据库连接 - PyMySQL 驱动：http://www.runoob.com/python3/python3-mysql.html
openpyxl 2.6.2：https://pypi.org/project/openpyxl/
"""
import pymysql
from openpyxl import Workbook

if __name__ == "__main__":
    wb = Workbook()
    ws = wb.active
    # ws["A1"] = 42
    # ws.append([1, 2, 3])
    # ws["A2"] = datetime.datetime.now()

    # 分别输入MySQL ip 地址，用户名，密码，数据库名称
    db = pymysql.connect("x.x.x.x", "user", "password", "database")
    cursor = db.cursor()
    cursor.execute("SELECT * FROM judge_item;")
    for data in cursor.fetchall():
        print(data)
        ws.append(data)

    # 保存到指定文件
    wb.save("C:\\Users\\12439\\Documents\\Projects\\hw2\\judge.xlsx")
    db.close()